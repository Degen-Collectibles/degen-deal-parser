import csv
import json
from io import BytesIO, StringIO
from inspect import signature
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook
from sqlmodel import Session, SQLModel, create_engine, select
from starlette.requests import Request

import app.cache as cache_module
from app.discord.bank_reconciliation import rerun_bank_reconciliation
from app.ledger import (
    LedgerFilters,
    apply_ledger_automation,
    apply_ledger_rule,
    build_ledger_page_data,
    draft_ledger_rule_from_instruction,
    ledger_filters_from_values,
    ledger_status_for_bank_row,
    preview_ledger_automation,
    preview_ledger_review_agent,
    preview_ledger_rule,
    run_ledger_review_agent,
)
from app.models import AuditLog, AvailableDiscordChannel, BankStatementImport, BankTransaction, LedgerRule, Transaction
from app.models import DiscordMessage, PARSE_PARSED, PARSE_REVIEW_REQUIRED
import app.routers.ledger as ledger_routes
from app.routers.ledger import (
    ledger_agent_run_form,
    ledger_automation_apply_form,
    ledger_export_csv,
    ledger_export_xlsx,
    ledger_page,
    ledger_row_force_unmatch_form,
    ledger_row_status_form,
)


def make_request(path: str, role: str = "admin", *, method: str = "GET", headers: list[tuple[bytes, bytes]] | None = None) -> Request:
    request = Request(
        {
            "type": "http",
            "method": method,
            "path": path,
            "headers": headers or [],
            "scheme": "http",
            "server": ("testserver", 80),
            "client": ("testclient", 50000),
            "root_path": "",
        }
    )
    request.state.current_user = SimpleNamespace(
        username="tester",
        display_name="Test Operator",
        role=role,
    )
    return request


def make_engine():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False})
    SQLModel.metadata.create_all(engine)
    return engine


def add_import(session: Session) -> BankStatementImport:
    row = BankStatementImport(
        label="Chase feed",
        account_label="Chase Checking",
        account_type="checking",
        source_kind="plaid",
        provider="plaid",
    )
    session.add(row)
    session.commit()
    session.refresh(row)
    return row


def test_ledger_builder_counts_bank_rows_and_separates_unbanked_cash():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)
    cash_at = datetime(2026, 5, 14, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        matched_tx = Transaction(
            id=500,
            source_message_id=1500,
            occurred_at=posted_at,
            parse_status="parsed",
            entry_kind="buy",
            payment_method="zelle",
            expense_category="inventory",
            amount=250.0,
            money_in=0.0,
            money_out=250.0,
            source_content="buy inventory 250 zelle",
        )
        cash_tx = Transaction(
            id=501,
            source_message_id=1501,
            occurred_at=cash_at,
            parse_status="parsed",
            entry_kind="buy",
            payment_method="cash",
            expense_category="inventory",
            amount=90.0,
            money_in=0.0,
            money_out=90.0,
            source_content="buy inventory 90 cash",
        )
        cash_app_tx = Transaction(
            id=503,
            source_message_id=1503,
            occurred_at=cash_at,
            parse_status="parsed",
            entry_kind="buy",
            payment_method="cash_app",
            expense_category="inventory",
            amount=55.0,
            money_in=0.0,
            money_out=55.0,
            source_content="buy inventory 55 cash app",
        )
        session.add(matched_tx)
        session.add(cash_tx)
        session.add(cash_app_tx)
        session.add(
            BankTransaction(
                id=10,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="Zelle payment to inventory seller",
                amount=-250.0,
                classification="logged_in_discord_strong",
                confidence="high",
                expense_category="inventory_purchases",
                matched_transaction_id=500,
                matched_source_message_id=1500,
                matched_platform="discord",
            )
        )
        session.add(
            BankTransaction(
                id=11,
                import_id=bank_import.id,
                row_index=2,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="SHOPIFY PAYOUT 123",
                amount=600.0,
                classification="shopify_payout",
                confidence="high",
                expense_category="platform_payouts",
                matched_platform="shopify",
            )
        )
        session.add(
            BankTransaction(
                id=12,
                import_id=bank_import.id,
                row_index=3,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="PYMT SENT APPLE CASH SENT MONEY CUPERTINO CA",
                amount=-280.0,
                classification="direct_payment_out_needs_log_check",
                confidence="medium",
                expense_category="inventory_purchases",
            )
        )
        session.commit()

        data = build_ledger_page_data(session, LedgerFilters(status="all"))
        with_cash = build_ledger_page_data(session, LedgerFilters(status="all", include_cash=True))
        needs_action_with_cash = build_ledger_page_data(session, LedgerFilters(status="needs_action", include_cash=True))

    rows_by_id = {row["id"]: row for row in data["rows"]}
    cash_rows = [row for row in with_cash["rows"] if row["row_kind"] == "cash"]
    assert data["summary"]["bank_row_count"] == 3
    assert data["summary"]["bank_net_total"] == 70.0
    assert rows_by_id[10]["source"] == "discord"
    assert rows_by_id[10]["ledger_status"] == "reconciled"
    assert rows_by_id[11]["source"] == "shopify"
    assert rows_by_id[11]["ledger_status"] == "reconciled"
    assert rows_by_id[12]["ledger_status"] == "needs_action"
    assert rows_by_id[12]["action_reason_label"] == "Needs match check"
    assert data["unbanked_cash_rows"][0]["transaction_id"] == 501
    assert data["summary"]["unbanked_cash_total"] == 90.0
    assert len(with_cash["rows"]) == 4
    assert cash_rows[0]["id"] == "cash-501"
    assert cash_rows[0]["source"] == "cash"
    assert cash_rows[0]["description"] == "buy inventory 90 cash"
    assert "cash-503" not in {row["id"] for row in with_cash["rows"]}
    assert "cash-501" not in {row["id"] for row in needs_action_with_cash["rows"]}
    assert with_cash["summary"]["bank_net_total"] == 70.0


def test_ledger_builder_labels_money_events_with_combined_evidence():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            Transaction(
                id=610,
                source_message_id=1610,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="zelle",
                expense_category="inventory",
                amount=250.0,
                money_in=0.0,
                money_out=250.0,
                source_content="buy inventory 250 zelle",
            )
        )
        session.add(
            BankTransaction(
                id=61,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="ZELLE PAYMENT TO CARD SELLER",
                amount=-250.0,
                classification="logged_in_discord_strong",
                confidence="high",
                expense_category="inventory_purchases",
                matched_transaction_id=610,
                matched_source_message_id=1610,
                matched_platform="discord",
            )
        )
        session.add(
            BankTransaction(
                id=62,
                import_id=bank_import.id,
                row_index=2,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="SHOPIFY PAYOUT 123",
                amount=600.0,
                classification="shopify_payout",
                confidence="high",
                expense_category="platform_payouts",
                matched_platform="shopify",
            )
        )
        session.commit()

        data = build_ledger_page_data(session, LedgerFilters(status="all"))

    rows_by_id = {row["id"]: row for row in data["rows"]}
    matched_row = rows_by_id[61]
    payout_row = rows_by_id[62]

    assert matched_row["event_label"] == "Cash movement"
    assert [chip["label"] for chip in matched_row["evidence_chips"]] == ["Bank", "Discord log"]
    assert "bank cash movement matched to the Discord deal log" in matched_row["accounting_note"]

    assert payout_row["event_label"] == "Cash settlement"
    assert [chip["label"] for chip in payout_row["evidence_chips"]] == ["Bank", "Shopify payout"]
    assert "Finance revenue uses paid Shopify order rows" in payout_row["accounting_note"]


def test_ledger_dedupes_relinked_plaid_account_rows_but_keeps_repeated_charges():
    engine = make_engine()
    posted_at = datetime(2026, 5, 25, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        first_import = BankStatementImport(
            label="Original Chase feed",
            account_label="Chase Ultimate Rewards 2024",
            account_type="credit_card",
            source_kind="plaid",
            provider="plaid",
            provider_account_id="acc_old",
        )
        second_import = BankStatementImport(
            label="Relinked Chase feed",
            account_label="Chase Ultimate Rewards 2024",
            account_type="credit_card",
            source_kind="plaid",
            provider="plaid",
            provider_account_id="acc_new",
        )
        session.add(first_import)
        session.add(second_import)
        session.commit()
        session.refresh(first_import)
        session.refresh(second_import)
        for import_id, provider_prefix in ((first_import.id, "old"), (second_import.id, "new")):
            for row_index in (1, 2):
                session.add(
                    BankTransaction(
                        import_id=import_id or 0,
                        row_index=row_index,
                        account_label="Chase Ultimate Rewards 2024",
                        account_type="credit_card",
                        posted_at=posted_at,
                        transaction_at=posted_at,
                        description="Stamps.com",
                        description_stem="STAMPS.COM",
                        amount=-10.0,
                        classification="expense_or_purchase_needs_review",
                        expense_category="shipping_postage",
                        provider_transaction_id=f"{provider_prefix}_{row_index}",
                    )
                )
        session.commit()

        data = build_ledger_page_data(session, LedgerFilters(status="all", source="bank"))

    assert data["summary"]["bank_row_count"] == 2
    assert data["summary"]["filtered_row_count"] == 2
    assert data["summary"]["bank_outflow_total"] == -20.0
    assert [row["description"] for row in data["rows"]] == ["Stamps.com", "Stamps.com"]


def test_ledger_filters_needs_action_rows_by_action_reason():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=14,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="Customer Zelle payment",
                amount=180.0,
                classification="direct_customer_payment_needs_log_check",
                expense_category="sales_collections",
            )
        )
        session.add(
            BankTransaction(
                id=15,
                import_id=bank_import.id,
                row_index=2,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="WWW.PSACARD.COM",
                amount=-140.0,
                classification="expense_or_purchase_needs_review",
                expense_category="uncategorized",
            )
        )
        session.add(
            BankTransaction(
                id=16,
                import_id=bank_import.id,
                row_index=3,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="PYMT SENT APPLE CASH SENT MONEY CUPERTINO CA",
                amount=-280.0,
                classification="direct_payment_out_needs_log_check",
                expense_category="inventory_purchases",
            )
        )
        session.commit()

        filters = LedgerFilters(status="needs_action")
        filters.action_reason = "expense_review"
        data = build_ledger_page_data(session, filters)

    assert [row["id"] for row in data["rows"]] == [15]
    assert data["rows"][0]["action_reason_label"] == "Expense review"
    assert {"value": "expense_review", "label": "Expense review"} in data["action_reason_choices"]


def test_ledger_review_agent_clears_false_discord_matches_and_auto_reviews_safe_expense():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            Transaction(
                id=520,
                source_message_id=1520,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="cash",
                expense_category="inventory",
                amount=11.99,
                money_in=0.0,
                money_out=11.99,
                source_content="Bought for 12 cash",
            )
        )
        session.add(
            Transaction(
                id=521,
                source_message_id=1521,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="unknown",
                expense_category="inventory",
                amount=2500.0,
                money_in=0.0,
                money_out=2500.0,
                source_content="Bought airbnb 2500$ (owe me)",
            )
        )
        session.add(
            BankTransaction(
                id=17,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="Amazon Prime Video",
                amount=-11.99,
                classification="logged_in_discord_possible",
                confidence="medium",
                expense_category="inventory_purchases",
                expense_subcategory="Matched app inventory transaction",
                category_confidence="high",
                category_reason="Matched a normalized Discord/app inventory buy.",
                matched_transaction_id=520,
                matched_source_message_id=1520,
                matched_platform="discord",
                raw_row_json=json.dumps({"Category": "Entertainment"}),
            )
        )
        session.add(
            BankTransaction(
                id=18,
                import_id=bank_import.id,
                row_index=2,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="PYMT SENT APPLE CASH BALANCE CUPERTINO CA",
                amount=-2500.0,
                classification="logged_in_discord_possible",
                confidence="medium",
                expense_category="inventory_purchases",
                matched_transaction_id=521,
                matched_source_message_id=1521,
                matched_platform="discord",
            )
        )
        session.commit()

        result = run_ledger_review_agent(session, filters=LedgerFilters(status="needs_action"), limit=50)
        amazon = session.get(BankTransaction, 17)
        apple_cash = session.get(BankTransaction, 18)

    assert result["updated_count"] == 2
    assert result["cleared_false_matches"] == 2
    assert result["auto_reviewed"] == 0
    assert amazon.matched_transaction_id is None
    assert amazon.matched_platform is None
    assert amazon.expense_category == "meals_entertainment"
    assert amazon.review_status == "open"
    assert ledger_status_for_bank_row(amazon) == "needs_action"
    assert apple_cash.matched_transaction_id is None
    assert apple_cash.matched_platform is None
    assert apple_cash.classification == "direct_payment_out_needs_log_check"
    assert apple_cash.review_status == "open"
    assert ledger_status_for_bank_row(apple_cash) == "needs_action"


def test_ledger_review_agent_leaves_medium_confidence_safe_category_open():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=22,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="Canva subscription",
                amount=-19.99,
                classification="expense_or_purchase_needs_review",
                confidence="low",
                expense_category="software_subscriptions",
                expense_subcategory="Software/subscription",
                category_confidence="medium",
                category_reason="Software or recurring subscription descriptor.",
            )
        )
        session.commit()

        result = run_ledger_review_agent(session, filters=LedgerFilters(status="needs_action"), limit=50)
        row = session.get(BankTransaction, 22)

    assert result["scanned_count"] == 1
    assert result["updated_count"] == 0
    assert result["auto_reviewed"] == 0
    assert row.review_status == "open"
    assert ledger_status_for_bank_row(row) == "needs_action"


def test_ledger_review_agent_can_process_prod_sized_safe_expense_batches():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        for index in range(1200):
            session.add(
                BankTransaction(
                    id=2000 + index,
                    import_id=bank_import.id,
                    row_index=index + 1,
                    account_label="Chase Checking",
                    account_type="checking",
                    posted_at=posted_at,
                    description=f"USPS POSTAGE #{index}",
                    amount=-5.0,
                    classification="expense_or_purchase_needs_review",
                    confidence="high",
                    expense_category="shipping_postage",
                    expense_subcategory="Shipping label/postage",
                    category_confidence="high",
                    category_reason="Carrier, postage, or shipping software descriptor.",
                )
            )
        session.commit()

        result = run_ledger_review_agent(session, filters=LedgerFilters(status="needs_action"), limit=1500)
        reviewed_count = len(
            session.exec(
                select(BankTransaction).where(BankTransaction.review_status == "reviewed")
            ).all()
        )

    assert result["scanned_count"] == 1200
    assert result["updated_count"] == 1200
    assert result["auto_reviewed"] == 1200
    assert reviewed_count == 1200


def test_preview_ledger_review_agent_reports_actions_without_mutating_rows():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            Transaction(
                id=552,
                source_message_id=2552,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="cash",
                expense_category="inventory",
                amount=11.99,
                money_in=0.0,
                money_out=11.99,
                source_content="Bought for 12 cash",
            )
        )
        session.add(
            BankTransaction(
                id=52,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="Amazon Prime Video",
                amount=-11.99,
                classification="logged_in_discord_possible",
                expense_category="inventory_purchases",
                category_confidence="high",
                matched_transaction_id=552,
                matched_source_message_id=2552,
                matched_platform="discord",
                raw_row_json=json.dumps({"Category": "Entertainment"}),
            )
        )
        session.add(
            BankTransaction(
                id=53,
                import_id=bank_import.id,
                row_index=2,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="USPS POSTAGE",
                amount=-8.25,
                classification="expense_or_purchase_needs_review",
                confidence="high",
                expense_category="shipping_postage",
                category_confidence="high",
            )
        )
        session.commit()

        preview = preview_ledger_review_agent(session, filters=LedgerFilters(status="needs_action"), limit=50)
        amazon = session.get(BankTransaction, 52)
        postage = session.get(BankTransaction, 53)
        audits = session.exec(select(AuditLog).where(AuditLog.action == "financial.ledger.agent_review")).all()

    assert preview["scanned_count"] == 2
    assert preview["updated_count"] == 2
    assert preview["cleared_false_matches"] == 1
    assert preview["auto_reviewed"] == 1
    assert {sample["id"] for sample in preview["sample_actions"]} == {52, 53}
    assert amazon.matched_transaction_id == 552
    assert amazon.matched_platform == "discord"
    assert amazon.review_status == "open"
    assert postage.review_status == "open"
    assert audits == []


def test_ledger_page_data_builds_tax_cleanup_summary():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=2400,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="ATM cash deposit",
                amount=500.0,
                classification="cash_deposit_needs_source",
                expense_category="cash_deposits",
                category_confidence="medium",
            )
        )
        session.add(
            BankTransaction(
                id=2401,
                import_id=bank_import.id,
                row_index=2,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="USPS POSTAGE",
                amount=-20.0,
                classification="expense_or_purchase_needs_review",
                expense_category="shipping_postage",
                category_confidence="high",
            )
        )
        session.add(
            BankTransaction(
                id=2402,
                import_id=bank_import.id,
                row_index=3,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="Amazon Prime Video",
                amount=-12.0,
                classification="logged_in_discord_possible",
                expense_category="meals_entertainment",
                category_confidence="medium",
            )
        )
        session.add(
            Transaction(
                id=2450,
                source_message_id=2450,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="cash",
                expense_category="inventory",
                amount=80.0,
                money_in=0.0,
                money_out=80.0,
                source_content="buy inventory 80 cash",
            )
        )
        session.commit()

        data = build_ledger_page_data(session, LedgerFilters(status="all", include_cash=True))

    cleanup = data["tax_cleanup"]
    buckets = {bucket["reason"]: bucket for bucket in cleanup["buckets"]}

    assert cleanup["status"] == "needs_cleanup"
    assert cleanup["needs_action_count"] == 3
    assert cleanup["agent_ready_count"] == 1
    assert cleanup["evidence_count"] == 2
    assert cleanup["cash_count"] == 1
    assert cleanup["cash_net_total"] == -80.0
    assert cleanup["cash_net_display"] == "-$80.00"
    assert buckets["needs_source"]["count"] == 1
    assert buckets["needs_source"]["exposure_total"] == 500.0
    assert buckets["expense_review"]["agent_ready_count"] == 1
    assert buckets["possible_discord_match"]["href"] == "/ledger?status=needs_action&action_reason=possible_discord_match"


def test_ledger_page_data_builds_quick_chip_urls_from_selected_filters():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=2451,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="SHOPIFY PAYOUT",
                amount=125.0,
                classification="shopify_payout",
                expense_category="platform_payouts",
            )
        )
        session.commit()

        data = build_ledger_page_data(
            session,
            LedgerFilters(
                start="2026-05-01",
                status="needs_action",
                search="apple",
                sort="amount",
                direction="asc",
                include_cash=False,
            ),
        )

    chips = {chip["label"]: chip for chip in data["quick_chips"]}
    assert chips["Shopify"]["href"] == (
        "/ledger?start=2026-05-01&status=all&source=shopify&search=apple&sort=amount"
        "&direction=asc&include_cash=false"
    )
    assert chips["Cash only"]["href"] == (
        "/ledger?start=2026-05-01&status=all&source=cash&search=apple&sort=amount"
        "&direction=asc&include_cash=true"
    )


def test_ledger_review_agent_preserves_manual_category_when_clearing_false_discord_match():
    _assert_ledger_review_agent_preserves_locked_category_when_clearing_false_discord_match("manual")


def test_ledger_review_agent_preserves_rule_category_when_clearing_false_discord_match():
    _assert_ledger_review_agent_preserves_locked_category_when_clearing_false_discord_match("rule")


def _assert_ledger_review_agent_preserves_locked_category_when_clearing_false_discord_match(category_confidence: str):
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)
    category_reason = {
        "manual": "Manually changed from the ledger.",
        "rule": "Ledger rule: PSA grading fees.",
    }[category_confidence]

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            Transaction(
                id=523,
                source_message_id=1523,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="cash",
                expense_category="inventory",
                amount=149.0,
                money_in=0.0,
                money_out=149.0,
                source_content="Bought for 149 cash",
            )
        )
        session.add(
            BankTransaction(
                id=23,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="WWW.PSACARD.COM",
                amount=-149.0,
                classification="logged_in_discord_possible",
                confidence="medium",
                expense_category="grading_fees",
                expense_subcategory="Manual override",
                category_confidence=category_confidence,
                category_reason=category_reason,
                matched_transaction_id=523,
                matched_source_message_id=1523,
                matched_platform="discord",
            )
        )
        session.commit()

        result = run_ledger_review_agent(session, filters=LedgerFilters(status="needs_action"), limit=50)
        row = session.get(BankTransaction, 23)

    assert result["updated_count"] == 1
    assert result["cleared_false_matches"] == 1
    assert row.matched_transaction_id is None
    assert row.matched_platform is None
    assert row.expense_category == "grading_fees"
    assert row.expense_subcategory == "Manual override"
    assert row.category_confidence == category_confidence
    assert row.category_reason == category_reason


def test_ledger_agent_route_requires_preview_before_apply():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            Transaction(
                id=522,
                source_message_id=1522,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="cash",
                expense_category="inventory",
                amount=11.99,
                money_in=0.0,
                money_out=11.99,
                source_content="Bought for 12 cash",
            )
        )
        session.add(
            BankTransaction(
                id=19,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="Amazon Prime Video",
                amount=-11.99,
                classification="logged_in_discord_possible",
                expense_category="inventory_purchases",
                category_confidence="high",
                matched_transaction_id=522,
                matched_source_message_id=1522,
                matched_platform="discord",
                raw_row_json=json.dumps({"Category": "Entertainment"}),
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_agent_run_form(
                make_request("/ledger/agent/run-form", method="POST"),
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="needs_action",
                selected_category="",
                selected_source="",
                selected_action_reason="possible_discord_match",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                selected_include_cash="",
                session=session,
            )
        row = session.get(BankTransaction, 19)
        audits = session.exec(select(AuditLog).where(AuditLog.action == "financial.ledger.agent_review")).all()

    assert response.status_code == 303
    assert "Preview+required" in response.headers["location"]
    assert row.matched_transaction_id == 522
    assert row.matched_platform == "discord"
    assert row.review_status == "open"
    assert audits == []


def test_ledger_agent_preview_route_renders_confirmation_without_mutating_rows():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            Transaction(
                id=524,
                source_message_id=1524,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="cash",
                expense_category="inventory",
                amount=11.99,
                money_in=0.0,
                money_out=11.99,
                source_content="Bought for 12 cash",
            )
        )
        session.add(
            BankTransaction(
                id=24,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="Amazon Prime Video",
                amount=-11.99,
                classification="logged_in_discord_possible",
                expense_category="inventory_purchases",
                category_confidence="high",
                matched_transaction_id=524,
                matched_source_message_id=1524,
                matched_platform="discord",
                raw_row_json=json.dumps({"Category": "Entertainment"}),
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_routes.ledger_agent_preview_form(
                make_request("/ledger/agent/preview-form", method="POST"),
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="needs_action",
                selected_category="",
                selected_source="",
                selected_action_reason="possible_discord_match",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                selected_include_cash="",
                session=session,
            )
        row = session.get(BankTransaction, 24)

    body = response.body.decode("utf-8")
    assert response.status_code == 200
    assert "Ledger Agent Preview" in body
    assert "This will update 1 row(s)" in body
    assert 'name="confirm" value="run_agent"' in body
    assert row.matched_transaction_id == 524
    assert row.matched_platform == "discord"


def test_ledger_agent_route_confirm_applies_and_redirects_to_selected_view():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            Transaction(
                id=525,
                source_message_id=1525,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="cash",
                expense_category="inventory",
                amount=11.99,
                money_in=0.0,
                money_out=11.99,
                source_content="Bought for 12 cash",
            )
        )
        session.add(
            BankTransaction(
                id=25,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="Amazon Prime Video",
                amount=-11.99,
                classification="logged_in_discord_possible",
                expense_category="inventory_purchases",
                category_confidence="high",
                matched_transaction_id=525,
                matched_source_message_id=1525,
                matched_platform="discord",
                raw_row_json=json.dumps({"Category": "Entertainment"}),
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_agent_run_form(
                make_request("/ledger/agent/run-form", method="POST"),
                confirm="run_agent",
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="needs_action",
                selected_category="",
                selected_source="",
                selected_action_reason="possible_discord_match",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                selected_include_cash="",
                session=session,
            )
        row = session.get(BankTransaction, 25)
        audits = session.exec(select(AuditLog).where(AuditLog.action == "financial.ledger.agent_review")).all()

    assert response.status_code == 303
    assert "action_reason=possible_discord_match" in response.headers["location"]
    assert "Ledger+agent+updated+1" in response.headers["location"]
    assert row.matched_transaction_id is None
    assert row.matched_platform is None
    assert row.review_status == "open"
    assert len(audits) == 1


def test_ledger_export_handles_cash_rows():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=13,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="ATM cash deposit",
                amount=200.0,
                classification="cash_deposit_needs_source",
                expense_category="cash_deposits",
            )
        )
        session.add(
            Transaction(
                id=504,
                source_message_id=1504,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="cash",
                expense_category="inventory",
                amount=90.0,
                money_in=0.0,
                money_out=90.0,
                source_content="buy inventory 90 cash",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_export_csv(
                make_request("/ledger/export.csv?status=all&include_cash=true"),
                account="",
                start="",
                end="",
                status="all",
                category="",
                source="",
                action_reason="",
                search="",
                sort="posted_at",
                direction="desc",
                include_cash=True,
                session=session,
            )

    body = response.body.decode("utf-8")
    assert response.status_code == 200
    assert "cash-504" in body
    assert "buy inventory 90 cash" in body


def test_ledger_export_xlsx_downloads_excel_with_editable_ledger_rows():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=13,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="ATM cash deposit",
                amount=200.0,
                classification="cash_deposit_needs_source",
                expense_category="cash_deposits",
                review_note="Needs owner source.",
            )
        )
        session.add(
            Transaction(
                id=504,
                source_message_id=1504,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="cash",
                expense_category="inventory",
                amount=90.0,
                money_in=0.0,
                money_out=90.0,
                source_content="buy inventory 90 cash",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_export_xlsx(
                make_request("/ledger/export.xlsx?status=all&include_cash=true"),
                account="",
                start="",
                end="",
                status="all",
                category="",
                source="",
                action_reason="",
                search="",
                sort="posted_at",
                direction="desc",
                include_cash=True,
                session=session,
            )

    assert response.status_code == 200
    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert 'filename="ledger-export.xlsx"' in response.headers["Content-Disposition"]

    workbook = load_workbook(BytesIO(response.body))
    sheet = workbook["Ledger"]
    headers = [cell.value for cell in sheet[1]]
    assert headers == [
        "row_id",
        "row_kind",
        "posted_at",
        "account",
        "amount",
        "ledger_status",
        "source",
        "category",
        "classification",
        "description",
        "matched_transaction_id",
        "match_reason",
        "review_status",
        "review_note",
    ]
    values_by_row_id = {sheet.cell(row=row_index, column=1).value: row_index for row_index in range(2, sheet.max_row + 1)}
    bank_row = values_by_row_id[13]
    cash_row = values_by_row_id["cash-504"]
    assert sheet.cell(row=bank_row, column=10).value == "ATM cash deposit"
    assert sheet.cell(row=bank_row, column=14).value == "Needs owner source."
    assert sheet.cell(row=cash_row, column=2).value == "cash"
    assert sheet.cell(row=cash_row, column=10).value == "buy inventory 90 cash"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == f"A1:N{sheet.max_row}"


def test_ledger_exports_escape_formula_like_text_cells():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)
    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=14,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description='=HYPERLINK("http://evil.test","x")',
                amount=-200.0,
                classification="expense_or_purchase_needs_review",
                expense_category="supplies_packaging",
                match_reason="+SUM(1,2)",
                review_note="@cmd",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            csv_response = ledger_export_csv(
                make_request("/ledger/export.csv?status=all&include_cash=false"),
                account="",
                start="",
                end="",
                status="all",
                category="",
                source="",
                action_reason="",
                search="",
                sort="posted_at",
                direction="desc",
                include_cash=False,
                session=session,
            )
            xlsx_response = ledger_export_xlsx(
                make_request("/ledger/export.xlsx?status=all&include_cash=false"),
                account="",
                start="",
                end="",
                status="all",
                category="",
                source="",
                action_reason="",
                search="",
                sort="posted_at",
                direction="desc",
                include_cash=False,
                session=session,
            )

    exported_row = next(csv.DictReader(StringIO(csv_response.body.decode("utf-8"))))
    assert exported_row["description"].startswith("'=")
    assert exported_row["match_reason"].startswith("'+")
    assert exported_row["review_note"].startswith("'@")

    workbook = load_workbook(BytesIO(xlsx_response.body), data_only=False)
    sheet = workbook["Ledger"]
    assert sheet.cell(row=2, column=10).value.startswith("'=")
    assert sheet.cell(row=2, column=10).data_type != "f"
    assert sheet.cell(row=2, column=12).value.startswith("'+")
    assert sheet.cell(row=2, column=14).value.startswith("'@")


def test_ledger_includes_non_cash_discord_financial_channel_rows():
    engine = make_engine()
    posted_at = datetime(2026, 5, 19, 12, tzinfo=timezone.utc)
    with Session(engine) as session:
        session.add(
            Transaction(
                id=610,
                source_message_id=1610,
                discord_message_id="financial-payroll",
                channel_id="financials-channel",
                channel_name="financials",
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="expense",
                payment_method="zelle",
                expense_category="payroll",
                amount=6500.0,
                money_in=0.0,
                money_out=6500.0,
                source_content="Pay Sam payroll of may 6500$",
            )
        )
        session.add(
            Transaction(
                id=611,
                source_message_id=1611,
                discord_message_id="store-zelle",
                channel_id="store-channel",
                channel_name="store-sales-and-trades",
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="sale",
                payment_method="zelle",
                expense_category="inventory",
                amount=100.0,
                money_in=100.0,
                money_out=0.0,
                source_content="sold slab 100 zelle",
            )
        )
        session.commit()

        data = build_ledger_page_data(session, LedgerFilters(status="all", source="discord", include_cash=True))

    ids = {row["id"] for row in data["rows"]}
    financial_row = next(row for row in data["rows"] if row["id"] == "discord-financial-610")

    assert "discord-financial-610" in ids
    assert "discord-financial-611" not in ids
    assert financial_row["row_kind"] == "discord_financial"
    assert financial_row["source"] == "discord"
    assert financial_row["amount"] == -6500.0
    assert financial_row["ledger_status"] == "reconciled"
    assert financial_row["matched_transaction_id"] == 610


def test_ledger_includes_non_cash_year_past_show_discord_deal_rows():
    engine = make_engine()
    posted_at = datetime(2026, 5, 9, 12, tzinfo=timezone.utc)
    with Session(engine) as session:
        session.add(
            AvailableDiscordChannel(
                channel_id="past-show-channel",
                channel_name="2026-may-9-eastbaycardshow",
                guild_id="1",
                guild_name="Degen Guild",
                category_name="2026 Past Shows",
                label="2026 Past Shows / #2026-may-9-eastbaycardshow",
            )
        )
        session.add(
            Transaction(
                id=612,
                source_message_id=1612,
                discord_message_id="past-show-zelle",
                channel_id="past-show-channel",
                channel_name="2026-may-9-eastbaycardshow",
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="sale",
                payment_method="zelle",
                expense_category="inventory",
                amount=225.0,
                money_in=225.0,
                money_out=0.0,
                source_content="sold slab 225 zelle",
            )
        )
        session.commit()

        data = build_ledger_page_data(session, LedgerFilters(status="all", source="discord", include_cash=True))

    deal_row = next(row for row in data["rows"] if row["id"] == "discord-deal-612")

    assert deal_row["row_kind"] == "discord_deal"
    assert deal_row["account_label"] == "Discord Deals"
    assert deal_row["source"] == "discord"
    assert deal_row["classification"] == "discord_deal_log"
    assert deal_row["amount"] == 225.0
    assert deal_row["ledger_status"] == "reconciled"
    assert deal_row["matched_transaction_id"] == 612


def test_ledger_includes_non_cash_offline_purchase_discord_deal_rows():
    engine = make_engine()
    posted_at = datetime(2026, 5, 25, 21, tzinfo=timezone.utc)
    with Session(engine) as session:
        session.add(
            AvailableDiscordChannel(
                channel_id="alex-purchases-channel",
                channel_name="alex-purchases",
                guild_id="1",
                guild_name="Degen Guild",
                category_name="Offline Deals",
                label="Offline Deals / #alex-purchases",
            )
        )
        session.add(
            Transaction(
                id=613,
                source_message_id=1613,
                discord_message_id="alex-purchase-zelle",
                channel_id="alex-purchases-channel",
                channel_name="alex-purchases",
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="zelle",
                expense_category="inventory",
                amount=500.0,
                money_in=0.0,
                money_out=500.0,
                source_content="Buy $500 zelle",
            )
        )
        session.commit()

        data = build_ledger_page_data(session, LedgerFilters(status="all", source="discord", include_cash=True))

    deal_row = next(row for row in data["rows"] if row["id"] == "discord-deal-613")

    assert deal_row["row_kind"] == "discord_deal"
    assert deal_row["account_label"] == "Discord Deals"
    assert deal_row["source"] == "discord"
    assert deal_row["classification"] == "discord_deal_log"
    assert deal_row["amount"] == -500.0
    assert deal_row["matched_transaction_id"] == 613


def test_ledger_page_links_discord_matches_to_deal_detail():
    engine = make_engine()
    posted_at = datetime(2026, 5, 19, 12, tzinfo=timezone.utc)
    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            Transaction(
                id=620,
                source_message_id=1620,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="zelle",
                expense_category="inventory",
                amount=250.0,
                money_in=0.0,
                money_out=250.0,
                source_content="buy cards 250 zelle",
            )
        )
        session.add(
            Transaction(
                id=621,
                source_message_id=1621,
                channel_name="financials",
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="expense",
                payment_method="zelle",
                expense_category="payroll",
                amount=650.0,
                money_in=0.0,
                money_out=650.0,
                source_content="payroll 650 zelle",
            )
        )
        session.add(
            BankTransaction(
                id=120,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="Zelle payment to seller",
                amount=-250.0,
                classification="logged_in_discord_strong",
                confidence="high",
                expense_category="inventory_purchases",
                matched_transaction_id=620,
                matched_source_message_id=1620,
                matched_platform="discord",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_page(
                make_request("/ledger?status=all&source=discord"),
                account="",
                start="",
                end="",
                status="all",
                category="",
                source="discord",
                action_reason="",
                search="",
                sort="posted_at",
                direction="desc",
                include_cash=True,
                session=session,
            )

    body = response.body.decode("utf-8")
    assert response.status_code == 200
    assert 'href="/deals/1620?return_path=%2Fledger"' in body
    assert 'href="/deals/1621?return_path=%2Fledger"' in body
    assert 'action="/ledger/transactions/1621/edit-form"' in body
    assert "Edit transaction" in body


def test_ledger_page_exposes_bulk_selectable_discord_deal_rows():
    engine = make_engine()
    posted_at = datetime(2026, 5, 19, 12, tzinfo=timezone.utc)
    with Session(engine) as session:
        session.add(
            AvailableDiscordChannel(
                channel_id="offline-purchases",
                channel_name="jeff-purchases",
                guild_id="1",
                guild_name="Degen Guild",
                category_name="Offline Deals",
                label="Offline Deals / #jeff-purchases",
            )
        )
        session.add(
            DiscordMessage(
                id=1630,
                discord_message_id="deal-message-1630",
                channel_id="offline-purchases",
                channel_name="jeff-purchases",
                author_name="tester",
                content="Buy $250 zelle",
                created_at=posted_at,
                parse_status=PARSE_PARSED,
                deal_type="buy",
                entry_kind="buy",
                payment_method="zelle",
                expense_category="inventory",
                amount=250.0,
                money_in=0.0,
                money_out=250.0,
                needs_review=False,
            )
        )
        session.add(
            Transaction(
                id=630,
                source_message_id=1630,
                discord_message_id="deal-message-1630",
                channel_id="offline-purchases",
                channel_name="jeff-purchases",
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="zelle",
                expense_category="inventory",
                amount=250.0,
                money_in=0.0,
                money_out=250.0,
                source_content="Buy $250 zelle",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_page(
                make_request("/ledger?status=all&source=discord"),
                account="",
                start="",
                end="",
                status="all",
                category="",
                source="discord",
                action_reason="",
                search="",
                sort="posted_at",
                direction="desc",
                include_cash=True,
                session=session,
            )

    body = response.body.decode("utf-8")
    assert response.status_code == 200
    assert 'class="row-select" type="checkbox" name="row_id" value="discord-deal-630"' in body
    assert 'data-row-edit-form data-ledger-transaction-form="discord-deal-630"' in body
    assert 'action="/ledger/transactions/1630/edit-form"' in body
    assert 'name="entry_kind"' in body
    assert '<option value="buy" selected>Buy</option>' in body
    assert 'name="amount" value="250.0"' in body
    assert 'name="payment_method"' in body
    assert '<option value="zelle" selected>Zelle</option>' in body
    assert 'name="expense_category"' in body
    assert 'data-discord-category-form="discord-deal-630"' not in body
    assert "Save category" not in body


def test_ledger_transaction_edit_form_can_return_json_for_in_place_updates():
    from app.routers import ledger as ledger_router

    edit_form = getattr(ledger_router, "ledger_transaction_edit_form", None)
    assert edit_form is not None

    engine = make_engine()
    occurred_at = datetime(2026, 5, 19, 12, tzinfo=timezone.utc)
    with Session(engine) as session:
        message = DiscordMessage(
            id=1803,
            discord_message_id="cash-row-message",
            channel_id="offline-cash-channel",
            channel_name="offline-cash",
            author_name="tester",
            content="Bought singles 90 cash",
            created_at=occurred_at,
            parse_status=PARSE_PARSED,
            deal_type="buy",
            entry_kind="buy",
            payment_method="cash",
            amount=90.0,
            money_in=0.0,
            money_out=90.0,
            expense_category="inventory",
            needs_review=False,
        )
        session.add(message)
        session.add(
            Transaction(
                id=1804,
                source_message_id=1803,
                occurred_at=occurred_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="cash",
                expense_category="inventory",
                amount=90.0,
                money_in=0.0,
                money_out=90.0,
                source_content="Bought singles 90 cash",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = edit_form(
                make_request(
                    "/ledger/transactions/1803/edit-form",
                    method="POST",
                    headers=[(b"x-requested-with", b"fetch")],
                ),
                source_message_id=1803,
                entry_kind="buy",
                amount="90",
                payment_method="cash",
                expense_category="inventory_purchases",
                notes="",
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="all",
                selected_category="",
                selected_source="cash",
                selected_action_reason="",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                selected_include_cash="true",
                session=session,
            )
        message = session.get(DiscordMessage, 1803)
        transaction = session.exec(select(Transaction).where(Transaction.source_message_id == 1803)).one()

    payload = json.loads(response.body)
    assert response.status_code == 200
    assert payload["ok"] is True
    assert payload["row"]["expense_category"] == "inventory_purchases"
    assert payload["row"]["expense_category_label"] == "Inventory purchases"
    assert message.expense_category == "inventory_purchases"
    assert transaction.expense_category == "inventory_purchases"


def test_ledger_transaction_edit_form_preserves_missing_financial_fields_on_partial_post():
    from app.routers import ledger as ledger_router

    edit_form = getattr(ledger_router, "ledger_transaction_edit_form", None)
    assert edit_form is not None

    engine = make_engine()
    occurred_at = datetime(2026, 5, 19, 12, tzinfo=timezone.utc)
    with Session(engine) as session:
        message = DiscordMessage(
            id=1805,
            discord_message_id="cash-row-message-partial",
            channel_id="offline-cash-channel",
            channel_name="offline-cash",
            author_name="tester",
            content="Bought singles 90 cash",
            created_at=occurred_at,
            parse_status=PARSE_PARSED,
            deal_type="buy",
            entry_kind="buy",
            payment_method="cash",
            amount=90.0,
            money_in=0.0,
            money_out=90.0,
            expense_category="inventory",
            needs_review=False,
        )
        session.add(message)
        session.add(
            Transaction(
                id=1806,
                source_message_id=1805,
                occurred_at=occurred_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="cash",
                expense_category="inventory",
                amount=90.0,
                money_in=0.0,
                money_out=90.0,
                source_content="Bought singles 90 cash",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = edit_form(
                make_request(
                    "/ledger/transactions/1805/edit-form",
                    method="POST",
                    headers=[(b"x-requested-with", b"fetch")],
                ),
                source_message_id=1805,
                entry_kind=None,
                amount=None,
                payment_method=None,
                expense_category="inventory_purchases",
                notes=None,
                selected_source="cash",
                session=session,
            )
        message = session.get(DiscordMessage, 1805)
        transaction = session.exec(select(Transaction).where(Transaction.source_message_id == 1805)).one()

    payload = json.loads(response.body)
    assert response.status_code == 200
    assert payload["ok"] is True
    assert message.entry_kind == "buy"
    assert message.amount == 90.0
    assert message.payment_method == "cash"
    assert message.money_out == 90.0
    assert message.expense_category == "inventory_purchases"
    assert transaction.entry_kind == "buy"
    assert transaction.amount == 90.0
    assert transaction.payment_method == "cash"
    assert transaction.expense_category == "inventory_purchases"


def test_ledger_transaction_edit_form_preserves_trade_cash_legs_on_partial_post():
    from app.routers import ledger as ledger_router

    edit_form = getattr(ledger_router, "ledger_transaction_edit_form", None)
    assert edit_form is not None

    engine = make_engine()
    occurred_at = datetime(2026, 5, 19, 12, tzinfo=timezone.utc)
    with Session(engine) as session:
        message = DiscordMessage(
            id=1807,
            discord_message_id="trade-row-message-partial",
            channel_id="offline-cash-channel",
            channel_name="offline-cash",
            author_name="tester",
            content="Trade top out bottom in plus 195 zelle",
            created_at=occurred_at,
            parse_status=PARSE_PARSED,
            deal_type="trade",
            entry_kind="trade",
            payment_method="zelle",
            cash_direction="to_store",
            amount=195.0,
            money_in=195.0,
            money_out=400.0,
            expense_category="inventory",
            needs_review=False,
        )
        session.add(message)
        session.add(
            Transaction(
                id=1808,
                source_message_id=1807,
                occurred_at=occurred_at,
                parse_status="parsed",
                entry_kind="trade",
                payment_method="zelle",
                cash_direction="to_store",
                expense_category="inventory",
                amount=195.0,
                money_in=195.0,
                money_out=400.0,
                source_content="Trade top out bottom in plus 195 zelle",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = edit_form(
                make_request(
                    "/ledger/transactions/1807/edit-form",
                    method="POST",
                    headers=[(b"x-requested-with", b"fetch")],
                ),
                source_message_id=1807,
                entry_kind=None,
                amount=None,
                payment_method=None,
                expense_category="inventory_purchases",
                notes=None,
                selected_source="cash",
                session=session,
            )
        message = session.get(DiscordMessage, 1807)
        transaction = session.exec(select(Transaction).where(Transaction.source_message_id == 1807)).one()

    payload = json.loads(response.body)
    assert response.status_code == 200
    assert payload["ok"] is True
    assert message.entry_kind == "trade"
    assert message.amount == 195.0
    assert message.money_in == 195.0
    assert message.money_out == 400.0
    assert transaction.entry_kind == "trade"
    assert transaction.amount == 195.0
    assert transaction.money_in == 195.0
    assert transaction.money_out == 400.0
    assert transaction.expense_category == "inventory_purchases"


def test_ledger_transaction_edit_form_updates_discord_source_transaction():
    from app.routers import ledger as ledger_router

    edit_form = getattr(ledger_router, "ledger_transaction_edit_form", None)
    assert edit_form is not None

    engine = make_engine()
    occurred_at = datetime(2026, 5, 19, 12, tzinfo=timezone.utc)
    with Session(engine) as session:
        message = DiscordMessage(
            id=1801,
            discord_message_id="financial-image-only",
            channel_id="financials-channel",
            channel_name="financials",
            author_name="tester",
            content="April check!",
            created_at=occurred_at,
            parse_status=PARSE_REVIEW_REQUIRED,
            deal_type="unknown",
            entry_kind="unknown",
            amount=None,
            money_in=0.0,
            money_out=0.0,
            expense_category="uncategorized",
            needs_review=True,
        )
        session.add(message)
        session.add(
            Transaction(
                id=1802,
                source_message_id=1801,
                occurred_at=occurred_at,
                parse_status=PARSE_REVIEW_REQUIRED,
                entry_kind="unknown",
                payment_method="unknown",
                expense_category="uncategorized",
                amount=None,
                money_in=0.0,
                money_out=0.0,
                source_content="April check!",
                needs_review=True,
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = edit_form(
                make_request("/ledger/transactions/1801/edit-form", method="POST"),
                source_message_id=1801,
                entry_kind="expense",
                amount="-1200",
                payment_method="check",
                expense_category="taxes_licenses",
                notes="corrected from check image",
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="all",
                selected_category="",
                selected_source="discord",
                selected_action_reason="",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                selected_include_cash="true",
                session=session,
            )

        session.refresh(message)
        tx = session.exec(select(Transaction).where(Transaction.source_message_id == message.id)).one()

    assert response.status_code == 303
    assert response.headers["location"].startswith("/ledger?")
    assert "source=discord" in response.headers["location"]
    assert message.parse_status == PARSE_PARSED
    assert message.needs_review is False
    assert message.deal_type == "expense"
    assert message.entry_kind == "expense"
    assert message.amount == 1200.0
    assert message.money_in == 0.0
    assert message.money_out == 1200.0
    assert message.payment_method == "check"
    assert message.expense_category == "taxes_licenses"
    assert tx.amount == 1200.0
    assert tx.money_out == 1200.0
    assert tx.expense_category == "taxes_licenses"


def test_ledger_export_includes_all_matching_rows_not_just_first_page():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)
    with Session(engine) as session:
        bank_import = add_import(session)
        for idx in range(1005):
            session.add(
                BankTransaction(
                    id=2000 + idx,
                    import_id=bank_import.id,
                    row_index=idx + 1,
                    account_label="Chase Checking",
                    account_type="checking",
                    posted_at=posted_at,
                    description=f"SUPPLIES ROW {idx}",
                    amount=-1.0,
                    classification="expense_or_purchase_needs_review",
                    expense_category="supplies_packaging",
                )
            )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_export_csv(
                make_request("/ledger/export.csv?status=all&include_cash=false"),
                account="",
                start="",
                end="",
                status="all",
                category="",
                source="",
                action_reason="",
                search="",
                sort="posted_at",
                direction="desc",
                include_cash=False,
                session=session,
            )

    body = response.body.decode("utf-8")
    assert body.count("SUPPLIES ROW") == 1005


def test_ledger_export_xlsx_includes_all_matching_rows_not_just_first_page():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)
    with Session(engine) as session:
        bank_import = add_import(session)
        for idx in range(1005):
            session.add(
                BankTransaction(
                    id=3000 + idx,
                    import_id=bank_import.id,
                    row_index=idx + 1,
                    account_label="Chase Checking",
                    account_type="checking",
                    posted_at=posted_at,
                    description=f"SUPPLIES ROW {idx}",
                    amount=-1.0,
                    classification="expense_or_purchase_needs_review",
                    expense_category="supplies_packaging",
                )
            )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_export_xlsx(
                make_request("/ledger/export.xlsx?status=all&include_cash=false"),
                account="",
                start="",
                end="",
                status="all",
                category="",
                source="",
                action_reason="",
                search="",
                sort="posted_at",
                direction="desc",
                include_cash=False,
                session=session,
            )

    workbook = load_workbook(BytesIO(response.body), read_only=True)
    sheet = workbook["Ledger"]
    descriptions = [row[9] for row in sheet.iter_rows(min_row=2, values_only=True)]
    assert len(descriptions) == 1005
    assert descriptions.count("SUPPLIES ROW 1004") == 1


def test_ledger_force_unmatch_writes_financial_audit_log():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)
    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=190,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="Zelle payment to inventory seller",
                amount=-250.0,
                classification="logged_in_discord_strong",
                confidence="high",
                expense_category="inventory_purchases",
                matched_transaction_id=500,
                matched_source_message_id=1500,
                matched_platform="discord",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            ledger_row_force_unmatch_form(
                make_request("/ledger/rows/190/force-unmatch-form", method="POST"),
                190,
                mode="force",
                note="Bad match",
                session=session,
            )

        audit = session.exec(select(AuditLog).where(AuditLog.action == "financial.ledger.force_unmatch")).one()
        payload = json.loads(audit.details_json)

    assert audit.resource_key == "bank_transactions:190"
    assert payload["before"]["matched_transaction_id"] == 500
    assert payload["after"]["match_override_status"] == "force_unmatched"


def test_rule_draft_preview_and_apply_updates_only_matching_rows():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=20,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="PYMT SENT APPLE CASH SENT MONEY CUPERTINO CA",
                amount=-280.0,
                classification="direct_payment_out_needs_log_check",
                expense_category="uncategorized",
            )
        )
        session.add(
            BankTransaction(
                id=21,
                import_id=bank_import.id,
                row_index=2,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="WWW.PSACARD.COM",
                amount=-140.0,
                classification="expense_or_purchase_needs_review",
                expense_category="uncategorized",
            )
        )
        session.commit()

        draft = draft_ledger_rule_from_instruction(
            "Always categorize Apple Cash sent as inventory purchases and mark reviewed"
        )
        preview = preview_ledger_rule(
            session,
            conditions=draft["conditions"],
            actions=draft["actions"],
            filters=LedgerFilters(status="all"),
        )
        rule = LedgerRule(
            name=draft["name"],
            conditions_json=json.dumps(draft["conditions"]),
            actions_json=json.dumps(draft["actions"]),
        )
        session.add(rule)
        session.commit()
        session.refresh(rule)
        applied = apply_ledger_rule(session, rule, filters=LedgerFilters(status="all"), applied_by="tester")
        apple = session.get(BankTransaction, 20)
        psa = session.get(BankTransaction, 21)

    assert preview["affected_count"] == 1
    assert preview["sample_rows"][0]["id"] == 20
    assert applied["updated_count"] == 1
    assert apple.expense_category == "inventory_purchases"
    assert apple.category_confidence == "rule"
    assert apple.review_status == "reviewed"
    assert psa.expense_category == "uncategorized"
    assert psa.review_status == "open"


def test_preview_ledger_automation_targets_needs_log_check_rows_only():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=22,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="QuickPay with Zelle payment from Customer",
                amount=180.0,
                classification="direct_customer_payment_needs_log_check",
                expense_category="sales_collections",
            )
        )
        session.add(
            BankTransaction(
                id=23,
                import_id=bank_import.id,
                row_index=2,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="PYMT SENT APPLE CASH SENT MONEY CUPERTINO CA",
                amount=-280.0,
                classification="direct_payment_out_needs_log_check",
                expense_category="inventory_purchases",
            )
        )
        session.commit()

        preview = preview_ledger_automation(
            session,
            action_key="mark_needs_log_checked",
            filters=LedgerFilters(status="needs_action"),
        )

    assert preview["action_key"] == "mark_needs_log_checked"
    assert preview["affected_count"] == 1
    assert preview["sample_rows"][0]["id"] == 22
    assert "mark 1 needs-log-check row" in preview["summary"].lower()


def test_apply_ledger_automation_marks_needs_log_check_rows_reviewed_with_note():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=25,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="QuickPay with Zelle payment from Customer",
                amount=180.0,
                classification="direct_customer_payment_needs_log_check",
                expense_category="sales_collections",
                review_note="Verified order #123.",
            )
        )
        session.add(
            BankTransaction(
                id=26,
                import_id=bank_import.id,
                row_index=2,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="WWW.PSACARD.COM",
                amount=-140.0,
                classification="expense_or_purchase_needs_review",
                expense_category="uncategorized",
            )
        )
        session.commit()

        result = apply_ledger_automation(
            session,
            action_key="mark_needs_log_checked",
            filters=LedgerFilters(status="needs_action"),
            applied_by="tester",
        )
        log_check = session.get(BankTransaction, 25)
        expense = session.get(BankTransaction, 26)

    assert result["matched_count"] == 1
    assert result["updated_count"] == 1
    assert log_check.review_status == "reviewed"
    assert log_check.review_note == "Verified order #123.\nLog checked from ledger automation workbench by tester."
    assert ledger_status_for_bank_row(log_check) == "reconciled"
    assert expense.review_status == "open"


def test_ledger_automation_apply_form_redirects_and_updates_matching_rows():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=27,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="QuickPay with Zelle payment from Customer",
                amount=180.0,
                classification="direct_customer_payment_needs_log_check",
                expense_category="sales_collections",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_automation_apply_form(
                make_request("/ledger/automation/mark_needs_log_checked/apply-form", method="POST"),
                action_key="mark_needs_log_checked",
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="needs_action",
                selected_category="",
                selected_source="",
                selected_action_reason="needs_log_check",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                selected_include_cash="",
                session=session,
            )
        row = session.get(BankTransaction, 27)

    assert response.status_code == 303
    assert "action_reason=needs_log_check" in response.headers["location"]
    assert "Automation+updated+1+of+1" in response.headers["location"]
    assert row.review_status == "reviewed"


def test_ledger_automation_apply_form_hides_unexpected_exception_details(caplog):
    engine = make_engine()
    with Session(engine) as session:
        with (
            patch("app.routers.ledger.require_role_response", return_value=None),
            patch(
                "app.routers.ledger.apply_ledger_automation",
                side_effect=RuntimeError("database password is super-secret"),
            ),
        ):
            response = ledger_automation_apply_form(
                make_request("/ledger/automation/mark_needs_log_checked/apply-form", method="POST"),
                action_key="mark_needs_log_checked",
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="needs_action",
                selected_category="",
                selected_source="",
                selected_action_reason="needs_log_check",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                selected_include_cash="",
                session=session,
            )

    location = response.headers["location"]
    assert response.status_code == 303
    assert "An+unexpected+error+occurred" in location
    assert "database" not in location
    assert "super-secret" not in location
    assert "ledger automation apply failed" in caplog.text


def test_ledger_warning_uses_defined_css_variable():
    template = Path("app/templates/ledger.html").read_text()
    assert "var(--warning)" not in template
    assert "var(--warn)" in template


def test_ledger_template_exposes_tax_cleanup_panel():
    template = Path("app/templates/ledger.html").read_text()
    assert "Tax Cleanup" in template
    assert "Preview Tax Agent" in template
    assert "tax_cleanup.buckets" in template
    assert "agent_ready_count" in template


def test_apply_ledger_rule_appends_note_to_existing_review_note():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=24,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="Canva subscription",
                amount=-19.99,
                classification="expense_or_purchase_needs_review",
                expense_category="uncategorized",
                review_note="Keep receipt in Drive.",
            )
        )
        rule = LedgerRule(
            name="Canva note",
            conditions_json=json.dumps({"description_contains": "canva"}),
            actions_json=json.dumps({"note": "Monthly design software."}),
        )
        session.add(rule)
        session.commit()
        session.refresh(rule)

        applied = apply_ledger_rule(session, rule, filters=LedgerFilters(status="all"), applied_by="tester")
        row = session.get(BankTransaction, 24)

    assert applied["updated_count"] == 1
    assert row.review_note == "Keep receipt in Drive.\nMonthly design software."


def test_rule_draft_can_force_unmatch_rows_from_discord():
    draft = draft_ledger_rule_from_instruction("Unmatch Apple Cash rows from Discord and keep them reviewed")

    assert draft["conditions"]["description_contains"] == "apple cash"
    assert draft["actions"]["match_override_status"] == "force_unmatched"
    assert draft["actions"]["review_status"] == "reviewed"


def test_force_unmatched_survives_bank_reconciliation_rerun():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            Transaction(
                id=700,
                source_message_id=1700,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="apple_cash",
                expense_category="inventory",
                amount=280.0,
                money_in=0.0,
                money_out=280.0,
                source_content="buy inventory 280 apple cash",
            )
        )
        session.add(
            BankTransaction(
                id=30,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="PYMT SENT APPLE CASH SENT MONEY CUPERTINO CA",
                amount=-280.0,
                classification="logged_in_discord_possible",
                expense_category="inventory_purchases",
                category_confidence="manual",
                matched_transaction_id=700,
                matched_source_message_id=1700,
                matched_platform="discord",
                match_override_status="force_unmatched",
                match_override_note="Cash deal was already handled outside the bank feed.",
                review_status="reviewed",
            )
        )
        session.commit()

        rerun_bank_reconciliation(session, bank_import.id)
        row = session.get(BankTransaction, 30)

    assert row.match_override_status == "force_unmatched"
    assert row.matched_transaction_id is None
    assert row.matched_source_message_id is None
    assert row.matched_platform is None
    assert row.review_status == "reviewed"
    assert row.category_confidence == "manual"
    assert "forced unmatched" in row.match_reason.lower()


def test_ledger_route_renders_default_needs_action_grid():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=40,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="PYMT SENT APPLE CASH SENT MONEY CUPERTINO CA",
                amount=-280.0,
                classification="direct_payment_out_needs_log_check",
                expense_category="inventory_purchases",
            )
        )
        session.add(
            Transaction(
                id=502,
                source_message_id=2502,
                occurred_at=posted_at,
                parse_status="parsed",
                entry_kind="buy",
                payment_method="cash",
                expense_category="inventory",
                amount=75.0,
                money_in=0.0,
                money_out=75.0,
                source_content="cash buy 75",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_page(
                make_request("/ledger"),
                account="",
                start="",
                end="",
                status="needs_action",
                category="",
                source="",
                action_reason="",
                search="",
                sort="posted_at",
                direction="desc",
                include_cash=False,
                session=session,
            )
            cash_response = ledger_page(
                make_request("/ledger?include_cash=true&status=all"),
                account="",
                start="",
                end="",
                status="all",
                category="",
                source="",
                action_reason="",
                search="",
                sort="posted_at",
                direction="desc",
                include_cash=True,
                session=session,
            )

    body = response.body.decode("utf-8")
    cash_body = cash_response.body.decode("utf-8")
    assert response.status_code == 200
    assert "Unified Ledger" in body
    assert "Automation Workbench" in body
    assert 'action="/ledger/automation/mark_needs_log_checked/apply-form"' in body
    assert "Preview: 0 row(s)" in body
    assert "No needs-log-check rows match the current filters." in body
    assert "Mark 0 reviewed" in body
    assert "disabled" in body
    assert "Scope:" in body
    assert "Review log-check rows" in body
    assert "Ledger Assistant" in body
    assert 'href="/ledger?status=all&amp;sort=posted_at&amp;direction=desc&amp;include_cash=true"' in body
    assert "All transactions" in body
    assert 'href="/ledger/export.xlsx?' in body
    assert "Export Excel" in body
    assert 'href="/ledger?status=needs_action&amp;action_reason=needs_match_check&amp;sort=posted_at&amp;direction=desc&amp;include_cash=false"' in body
    assert 'name="action_reason"' in body
    assert 'action="/ledger/agent/preview-form"' in body
    assert "Preview Ledger Agent" in body
    assert "Needs action means rows that still need a category" in body
    assert "No, side panel only" in body
    assert "Yes, include rows" in body
    assert "PYMT SENT APPLE CASH" in body
    assert "Needs match check" in body
    assert 'data-ledger-row-id="cash-502"' not in body
    assert "cash buy 75" in cash_body
    assert 'data-ledger-row-id="cash-502"' in cash_body
    assert "Cash only" in cash_body


def test_ledger_defaults_to_all_transactions_with_discord_cash_in_grid():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)
    cash_at = datetime(2026, 5, 14, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=91,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="SHOPIFY PAYOUT 123",
                amount=125.0,
                classification="shopify_payout",
                confidence="high",
                expense_category="platform_payouts",
            )
        )
        session.add(
            Transaction(
                id=591,
                source_message_id=2591,
                occurred_at=cash_at,
                parse_status="parsed",
                entry_kind="sale",
                payment_method="cash",
                expense_category="inventory",
                amount=40.0,
                money_in=40.0,
                money_out=0.0,
                source_content="sold pokemon singles 40 cash",
            )
        )
        session.commit()

        filters = ledger_filters_from_values()
        data = build_ledger_page_data(session, filters)

    route_params = signature(ledger_page).parameters
    assert route_params["status"].default.default == "all"
    assert route_params["include_cash"].default.default is True
    assert filters.status == "all"
    assert filters.include_cash is True
    assert "SHOPIFY PAYOUT 123" in {row["description"] for row in data["rows"]}
    assert "sold pokemon singles 40 cash" in {row["description"] for row in data["rows"]}
    assert "cash-591" in {row["id"] for row in data["rows"]}


def test_ledger_template_uses_dense_full_width_review_surface():
    source = open("app/templates/ledger.html", encoding="utf-8").read()

    assert "min-width: 1160px" not in source
    assert "minmax(340px, 420px)" not in source
    assert 'class="ledger-shell"' in source
    assert 'class="quick-chip' in source
    assert 'id="ledger-tools-drawer"' in source
    assert "data-ledger-row-id" in source
    assert "data-row-edit-form" in source
    assert "data-action-reason" in source
    assert "document.addEventListener(\"keydown\"" in source
    assert "focusSearch" in source


def test_ledger_template_presents_unified_money_event_language():
    source = open("app/templates/ledger.html", encoding="utf-8").read()

    assert "Money Events" in source
    assert "Cash / Log" in source
    assert 'data-label="Evidence"' in source
    assert "row.evidence_chips" in source
    assert "Finance revenue uses paid order rows" in source
    assert "orders stay as context" not in source.lower()


def test_ledger_template_has_mobile_card_layout_and_page_scroll():
    source = open("app/templates/ledger.html", encoding="utf-8").read()

    assert "@media (max-width: 700px)" in source
    assert ".sheet-wrap {" in source
    assert "max-height: none" in source
    assert "overflow-y: visible" in source
    assert "tbody tr.ledger-row" in source
    assert "display: grid" in source
    assert "td::before" in source
    assert 'data-label="Status"' in source
    assert 'data-label="Description"' in source
    assert 'data-label="Review"' in source
    assert "touch-action: manipulation" in source
    assert "overflow-x: auto" not in source


def test_ledger_row_status_form_can_return_json_for_in_place_updates():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=50,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="PYMT SENT APPLE CASH SENT MONEY CUPERTINO CA",
                amount=-280.0,
                classification="direct_payment_out_needs_log_check",
                expense_category="uncategorized",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_row_status_form(
                make_request(
                    "/ledger/rows/50/status-form",
                    method="POST",
                    headers=[(b"x-requested-with", b"fetch")],
                ),
                row_id=50,
                review_status="reviewed",
                classification="",
                expense_category="inventory_purchases",
                note="handled in-grid",
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="needs_action",
                selected_category="",
                selected_source="",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                session=session,
            )
        row = session.get(BankTransaction, 50)

    assert response.status_code == 200
    assert json.loads(response.body)["ok"] is True
    assert json.loads(response.body)["row"]["review_status"] == "reviewed"
    assert json.loads(response.body)["row"]["action_reason_label"] == ""
    assert row.review_status == "reviewed"
    assert row.expense_category == "inventory_purchases"


def test_ledger_row_status_form_redirect_preserves_cash_hidden_filter():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=54,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="WWW.PSACARD.COM",
                amount=-140.0,
                classification="expense_or_purchase_needs_review",
                expense_category="uncategorized",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_row_status_form(
                make_request("/ledger/rows/54/status-form", method="POST"),
                row_id=54,
                review_status="reviewed",
                classification="",
                expense_category="grading_fees",
                note="checked",
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="all",
                selected_category="",
                selected_source="",
                selected_action_reason="",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                selected_include_cash="false",
                session=session,
            )

    assert response.status_code == 303
    location = response.headers["location"]
    assert "include_cash=false" in location
    assert "include_cash=true" not in location


def test_ledger_row_status_form_preserves_existing_review_note_when_note_is_blank():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=51,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="WWW.PSACARD.COM",
                amount=-140.0,
                classification="expense_or_purchase_needs_review",
                expense_category="uncategorized",
                review_note="Existing reviewer context.",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_row_status_form(
                make_request(
                    "/ledger/rows/51/status-form",
                    method="POST",
                    headers=[(b"x-requested-with", b"fetch")],
                ),
                row_id=51,
                review_status="reviewed",
                classification="",
                expense_category="grading_fees",
                note="",
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="needs_action",
                selected_category="",
                selected_source="",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                session=session,
            )
        row = session.get(BankTransaction, 51)

    assert response.status_code == 200
    assert row.review_status == "reviewed"
    assert row.expense_category == "grading_fees"
    assert row.review_note == "Existing reviewer context."


def test_ledger_row_status_form_preserves_existing_review_note_when_note_is_whitespace_only():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=52,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="WWW.PSACARD.COM",
                amount=-140.0,
                classification="expense_or_purchase_needs_review",
                expense_category="uncategorized",
                review_note="Existing reviewer context.",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_row_status_form(
                make_request(
                    "/ledger/rows/52/status-form",
                    method="POST",
                    headers=[(b"x-requested-with", b"fetch")],
                ),
                row_id=52,
                review_status="reviewed",
                classification="",
                expense_category="grading_fees",
                note="   \n\t  ",
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="needs_action",
                selected_category="",
                selected_source="",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                session=session,
            )
        row = session.get(BankTransaction, 52)

    assert response.status_code == 200
    assert row.review_status == "reviewed"
    assert row.expense_category == "grading_fees"
    assert row.review_note == "Existing reviewer context."


def test_ledger_row_status_form_preserves_existing_review_note_when_note_is_omitted():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=53,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="WWW.PSACARD.COM",
                amount=-140.0,
                classification="expense_or_purchase_needs_review",
                expense_category="uncategorized",
                review_note="Existing reviewer context.",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_row_status_form(
                make_request(
                    "/ledger/rows/53/status-form",
                    method="POST",
                    headers=[(b"x-requested-with", b"fetch")],
                ),
                row_id=53,
                review_status="reviewed",
                classification="",
                expense_category="grading_fees",
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="needs_action",
                selected_category="",
                selected_source="",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                session=session,
            )
        row = session.get(BankTransaction, 53)

    assert response.status_code == 200
    assert row.review_status == "reviewed"
    assert row.expense_category == "grading_fees"
    assert row.review_note == "Existing reviewer context."


def test_ledger_row_status_form_invalidates_report_and_finance_caches():
    cache_module._cache.clear()
    cache_module.cache_set("reports:test", {"stale": True})
    cache_module.cache_set("finance:v4:test", {"stale": True})
    cache_module.cache_set("other:test", {"keep": True})

    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    try:
        with Session(engine) as session:
            bank_import = add_import(session)
            session.add(
                BankTransaction(
                    id=55,
                    import_id=bank_import.id,
                    row_index=1,
                    account_label="Chase Checking",
                    account_type="checking",
                    posted_at=posted_at,
                    description="WWW.PSACARD.COM",
                    amount=-140.0,
                    classification="expense_or_purchase_needs_review",
                    expense_category="uncategorized",
                )
            )
            session.commit()

            with patch("app.routers.ledger.require_role_response", return_value=None):
                ledger_row_status_form(
                    make_request(
                        "/ledger/rows/55/status-form",
                        method="POST",
                        headers=[(b"x-requested-with", b"fetch")],
                    ),
                    row_id=55,
                    review_status="reviewed",
                    classification="",
                    expense_category="grading_fees",
                    selected_account="",
                    selected_start="",
                    selected_end="",
                    selected_status="needs_action",
                    selected_category="",
                    selected_source="",
                    selected_search="",
                    selected_sort="posted_at",
                    selected_direction="desc",
                    session=session,
                )
    finally:
        engine.dispose()

    assert cache_module.cache_get("reports:test") is None
    assert cache_module.cache_get("finance:v4:test") is None
    assert cache_module.cache_get("other:test") == {"keep": True}
    cache_module._cache.clear()


def test_ledger_row_status_form_updates_review_note_to_stripped_nonblank_note():
    engine = make_engine()
    posted_at = datetime(2026, 5, 15, 12, tzinfo=timezone.utc)

    with Session(engine) as session:
        bank_import = add_import(session)
        session.add(
            BankTransaction(
                id=54,
                import_id=bank_import.id,
                row_index=1,
                account_label="Chase Checking",
                account_type="checking",
                posted_at=posted_at,
                description="WWW.PSACARD.COM",
                amount=-140.0,
                classification="expense_or_purchase_needs_review",
                expense_category="uncategorized",
                review_note="Existing reviewer context.",
            )
        )
        session.commit()

        with patch("app.routers.ledger.require_role_response", return_value=None):
            response = ledger_row_status_form(
                make_request(
                    "/ledger/rows/54/status-form",
                    method="POST",
                    headers=[(b"x-requested-with", b"fetch")],
                ),
                row_id=54,
                review_status="reviewed",
                classification="",
                expense_category="grading_fees",
                note="  Updated reviewer context.  \n",
                selected_account="",
                selected_start="",
                selected_end="",
                selected_status="needs_action",
                selected_category="",
                selected_source="",
                selected_search="",
                selected_sort="posted_at",
                selected_direction="desc",
                session=session,
            )
        row = session.get(BankTransaction, 54)

    assert response.status_code == 200
    assert row.review_status == "reviewed"
    assert row.expense_category == "grading_fees"
    assert row.review_note == "Updated reviewer context."
