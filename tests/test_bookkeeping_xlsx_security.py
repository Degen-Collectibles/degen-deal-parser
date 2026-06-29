from __future__ import annotations

import asyncio
import re
import struct
import time
import zipfile
from contextlib import contextmanager
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import Workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel, Session, create_engine, select

from app.discord import bookkeeping as bookkeeping_module
from app.discord.bookkeeping import (
    import_bookkeeping_file,
    read_tabular_rows,
    refresh_bookkeeping_import_from_source,
)
from app.models import BookkeepingEntry, BookkeepingImport


def _xlsx_bytes(*, rows: int = 1, columns: int = 2, sheets: int = 1) -> bytes:
    workbook = Workbook()
    for sheet_index in range(sheets):
        sheet = workbook.active if sheet_index == 0 else workbook.create_sheet()
        sheet.title = f"Sheet {sheet_index + 1}"
        sheet.append([f"column_{index + 1}" for index in range(columns)])
        for row_index in range(rows):
            sheet.append([f"r{row_index}c{column_index}" for column_index in range(columns)])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _xlsx_with_declared_expanded_member_size(size: int) -> bytes:
    """Change only a central-directory size; the parser must reject before extraction."""

    payload = bytearray(_xlsx_bytes())
    central_header = payload.find(b"PK\x01\x02")
    assert central_header >= 0
    struct.pack_into("<I", payload, central_header + 24, size)
    return bytes(payload)


def _zip_with_entries(count: int) -> bytes:
    output = BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for index in range(count):
            archive.writestr(f"entry-{index}.xml", b"")
    return output.getvalue()


def _rewrite_first_sheet_dimension(content: bytes, dimension: str) -> bytes:
    source_buffer = BytesIO(content)
    output = BytesIO()
    with (
        zipfile.ZipFile(source_buffer, "r") as source,
        zipfile.ZipFile(output, "w") as rewritten,
    ):
        for member in source.infolist():
            member_content = source.read(member)
            if member.filename == "xl/worksheets/sheet1.xml":
                member_content, replacements = re.subn(
                    br'<dimension ref="[^"]+"\s*/>',
                    f'<dimension ref="{dimension}"/>'.encode("ascii"),
                    member_content,
                    count=1,
                )
                assert replacements == 1
            rewritten.writestr(member, member_content)
    return output.getvalue()


class TestBookkeepingXlsxSecurityLimits:
    def test_rejects_compressed_upload_over_eight_mib_before_zip_or_workbook_parse(self):
        content = b"x" * ((8 * 1024 * 1024) + 1)

        with (
            patch.object(bookkeeping_module.zipfile, "ZipFile") as zip_file,
            patch.object(bookkeeping_module, "load_workbook") as load_workbook,
            pytest.raises(
                bookkeeping_module.BookkeepingFileValidationError,
                match=f"^{bookkeeping_module.BOOKKEEPING_XLSX_VALIDATION_ERROR}$",
            ),
        ):
            read_tabular_rows("oversized.xlsx", content)

        zip_file.assert_not_called()
        load_workbook.assert_not_called()

    def test_rejects_more_than_256_central_directory_entries(self):
        content = _zip_with_entries(257)

        with (
            patch.object(bookkeeping_module, "load_workbook") as load_workbook,
            pytest.raises(bookkeeping_module.BookkeepingFileValidationError),
        ):
            read_tabular_rows("too-many-entries.xlsx", content)

        load_workbook.assert_not_called()

    def test_rejects_declared_expanded_size_over_sixteen_mib(self):
        content = _xlsx_with_declared_expanded_member_size((16 * 1024 * 1024) + 1)

        with (
            patch.object(bookkeeping_module, "load_workbook") as load_workbook,
            pytest.raises(bookkeeping_module.BookkeepingFileValidationError),
        ):
            read_tabular_rows("expanded-bomb.xlsx", content)

        load_workbook.assert_not_called()

    def test_malformed_xlsx_uses_stable_user_safe_error(self):
        with pytest.raises(
            bookkeeping_module.BookkeepingFileValidationError,
            match=f"^{bookkeeping_module.BOOKKEEPING_XLSX_VALIDATION_ERROR}$",
        ):
            read_tabular_rows("malformed.xlsx", b"not an xlsx archive")

    def test_rejects_more_than_sixteen_worksheets_and_closes_workbook(self):
        workbook = MagicMock()
        workbook.worksheets = [MagicMock() for _ in range(17)]
        workbook.sheetnames = [f"Sheet {index}" for index in range(17)]

        with (
            patch.object(bookkeeping_module, "load_workbook", return_value=workbook),
            pytest.raises(bookkeeping_module.BookkeepingFileValidationError),
        ):
            read_tabular_rows("too-many-sheets.xlsx", _xlsx_bytes())

        workbook.close.assert_called_once_with()

    def test_chart_sheets_count_toward_the_sixteen_sheet_limit(self):
        workbook = MagicMock()
        workbook.worksheets = []
        workbook.sheetnames = [f"Chart {index}" for index in range(17)]

        with (
            patch.object(bookkeeping_module, "load_workbook", return_value=workbook),
            pytest.raises(bookkeeping_module.BookkeepingFileValidationError),
        ):
            read_tabular_rows("too-many-chart-sheets.xlsx", _xlsx_bytes())

        workbook.close.assert_called_once_with()

    def test_rejects_sheet_wider_than_sixty_four_columns(self):
        with pytest.raises(bookkeeping_module.BookkeepingFileValidationError):
            read_tabular_rows("too-wide.xlsx", _xlsx_bytes(columns=65))

    def test_spoofed_low_dimension_cannot_hide_more_than_sixty_four_columns(self):
        content = _rewrite_first_sheet_dimension(
            _xlsx_bytes(rows=1, columns=65),
            "A1",
        )

        with (
            patch.object(bookkeeping_module, "load_workbook") as load_workbook,
            pytest.raises(bookkeeping_module.BookkeepingFileValidationError),
        ):
            read_tabular_rows("spoofed-dimension.xlsx", content)

        load_workbook.assert_not_called()

    def test_spoofed_low_dimension_does_not_truncate_valid_cells(self):
        content = _rewrite_first_sheet_dimension(
            _xlsx_bytes(rows=1, columns=2),
            "A1",
        )

        assert read_tabular_rows("spoofed-dimension.xlsx", content) == [
            {
                "__sheet_name": "Sheet 1",
                "column_1": "r0c0",
                "column_2": "r0c1",
            }
        ]

    def test_rejects_sparse_sheet_with_oversized_declared_row_dimension(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "amount"
        sheet["A1000000"] = 50
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        with pytest.raises(bookkeeping_module.BookkeepingFileValidationError):
            read_tabular_rows("oversized-dimension.xlsx", output.getvalue())

    def test_rejects_more_than_ten_thousand_data_rows(self):
        with pytest.raises(bookkeeping_module.BookkeepingFileValidationError):
            read_tabular_rows(
                "too-many-rows.xlsx",
                _xlsx_bytes(rows=5_001, columns=1, sheets=2),
            )

    def test_rejects_more_than_250000_visited_cells(self):
        with pytest.raises(bookkeeping_module.BookkeepingFileValidationError):
            read_tabular_rows("too-many-cells.xlsx", _xlsx_bytes(rows=4_000, columns=64))

    def test_rejects_when_streaming_deadline_expires_and_closes_workbook(self):
        sheet = MagicMock()
        sheet.title = "Sheet 1"
        sheet.max_column = 2
        sheet.max_row = 2
        sheet.iter_rows.return_value = iter([("date", "amount"), ("2026-01-01", 50)])
        workbook = MagicMock()
        workbook.worksheets = [sheet]

        with (
            patch.object(bookkeeping_module, "_validate_xlsx_archive"),
            patch.object(bookkeeping_module, "load_workbook", return_value=workbook),
            patch.object(time, "monotonic", side_effect=[10.0, 10.0, 16.0]),
            pytest.raises(bookkeeping_module.BookkeepingFileValidationError),
        ):
            read_tabular_rows("slow.xlsx", _xlsx_bytes())

        workbook.close.assert_called_once_with()

    def test_rechecks_deadline_after_load_with_no_worksheets(self):
        workbook = MagicMock()
        workbook.worksheets = []

        with (
            patch.object(bookkeeping_module, "_validate_xlsx_archive"),
            patch.object(bookkeeping_module, "load_workbook", return_value=workbook),
            patch.object(time, "monotonic", side_effect=[10.0, 16.0]),
            pytest.raises(bookkeeping_module.BookkeepingFileValidationError),
        ):
            read_tabular_rows("slow-empty.xlsx", _xlsx_bytes())

        workbook.close.assert_called_once_with()

    def test_uses_read_only_safe_openpyxl_options_and_keeps_valid_xlsx_behavior(self):
        content = _xlsx_bytes(rows=1, columns=2)
        real_load_workbook = bookkeeping_module.load_workbook

        with patch.object(bookkeeping_module, "load_workbook", wraps=real_load_workbook) as loader:
            rows = read_tabular_rows("valid.xlsx", content)

        assert rows == [
            {
                "__sheet_name": "Sheet 1",
                "column_1": "r0c0",
                "column_2": "r0c1",
            }
        ]
        _, kwargs = loader.call_args
        assert kwargs == {"read_only": True, "data_only": True, "keep_links": False}

    def test_csv_behavior_is_unchanged(self):
        assert read_tabular_rows(
            "valid.csv", b"date,kind,amount\n2026-01-01,sale,50\n"
        ) == [
            {
                "__sheet_name": "import",
                "date": "2026-01-01",
                "kind": "sale",
                "amount": "50",
            }
        ]

    def test_rejection_cannot_partially_persist_an_import(self):
        engine = create_engine(
            "sqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        SQLModel.metadata.create_all(engine)

        with Session(engine) as session:
            with pytest.raises(bookkeeping_module.BookkeepingFileValidationError):
                import_bookkeeping_file(
                    session,
                    filename="too-wide.xlsx",
                    content=_xlsx_bytes(columns=65),
                    show_label="Security test",
                    show_date=None,
                    range_start=None,
                    range_end=None,
                )

            assert session.exec(select(BookkeepingImport)).all() == []

    def test_refresh_rejection_preserves_existing_import_and_entries(self):
        engine = create_engine(
            "sqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        SQLModel.metadata.create_all(engine)
        with Session(engine) as session:
            existing = BookkeepingImport(
                show_label="Existing import",
                source_kind="google_sheet_auto",
                source_name="existing.xlsx",
                source_url="https://docs.google.com/spreadsheets/d/existing/edit",
                row_count=1,
            )
            session.add(existing)
            session.flush()
            existing_id = existing.id
            session.add(
                BookkeepingEntry(
                    import_id=existing_id,
                    row_index=1,
                    amount=50,
                    raw_row_json='{"amount": 50}',
                )
            )
            session.commit()

        @contextmanager
        def test_session():
            with Session(engine) as session:
                yield session

        with (
            patch.object(bookkeeping_module, "managed_session", test_session),
            patch.object(
                bookkeeping_module,
                "fetch_google_sheet_export",
                new=AsyncMock(return_value=b"not an xlsx archive"),
            ),
            pytest.raises(bookkeeping_module.BookkeepingFileValidationError),
        ):
            asyncio.run(refresh_bookkeeping_import_from_source(existing_id))

        with Session(engine) as session:
            assert session.get(BookkeepingImport, existing_id) is not None
            assert len(session.exec(select(BookkeepingEntry)).all()) == 1
